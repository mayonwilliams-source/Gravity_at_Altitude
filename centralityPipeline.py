"""
Gravity at Altitude — H3 Network Centrality Pipeline
=====================================================
Pan-Himalayan Airlines (H3) | ISyE 6739

Pipeline:
  1. Load airport parameters from Excel
  2. Build Haversine distance matrix
  3. Build gravity-model transition matrix for each k value
  4. Compute Markov chain steady-state (centrality) via power iteration
  5. Sensitivity analysis across k = {0.0, 0.5, 0.75, 1.0, 1.5}
  6. Spearman rank correlation (centrality vs. raw size) per k
  7. 13x13 illustrative submatrix for appendix
  8. Save all results to Excel

Gravity model (fitted via MLE/OLS in JMP on 60 calibration routes):
  D(i,j) = C * size_i^alpha_o * size_j^alpha_d / distance(i,j)^k
  C = 37.746, alpha_o = 0.1166, alpha_d = 0.1017
  Distance term dropped from MLE fit (p=0.267, per-flight confound);
  k treated as sensitivity parameter with theoretical motivation.

Requirements: numpy, pandas, openpyxl, scipy
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import spearmanr

# ── 1. Load airport parameters ────────────────────────────────────────────────
wb = load_workbook("H3_data_collection_template-1.xlsx", data_only=True)
ws = wb["Airport Parameters"]

airports = []
for r in range(2, 132):
    iata = ws.cell(row=r, column=1).value
    if iata is None:
        continue
    airports.append(dict(
        iata  = iata,
        name  = ws.cell(row=r, column=2).value,
        lat   = float(ws.cell(row=r, column=3).value),
        lon   = float(ws.cell(row=r, column=4).value),
        h3pax = ws.cell(row=r, column=5).value or 0,
        share = ws.cell(row=r, column=6).value or 0,
        total = float(ws.cell(row=r, column=7).value),
        rank  = ws.cell(row=r, column=8).value,
    ))

df   = pd.DataFrame(airports).reset_index(drop=True)
N    = len(df)
size = df["total"].values
lat  = df["lat"].values
lon  = df["lon"].values

print(f"Loaded {N} airports")

# ── 2. Gravity model parameters (from JMP) ───────────────────────────────────
C       = 37.746   # scale constant  (exp of OLS intercept)
alpha_o = 0.1166   # origin size exponent
alpha_d = 0.1017   # destination size exponent
# Note: distance decay exponent k is varied in sensitivity analysis below.
# k was statistically insignificant in OLS fit (p=0.267) due to per-flight
# confound; treated here as a theoretically-motivated sensitivity parameter.

# ── 3. Haversine distance matrix (nautical miles) ────────────────────────────
def haversine_matrix(lat, lon):
    """Full N×N great-circle distance matrix in nautical miles."""
    R = 3440.065  # Earth radius in nm
    la = np.radians(lat)[:, None]
    lo = np.radians(lon)[:, None]
    dlat = la - la.T
    dlon = lo - lo.T
    a = np.sin(dlat/2)**2 + np.cos(la) * np.cos(la.T) * np.sin(dlon/2)**2
    d = 2 * R * np.arcsin(np.sqrt(np.clip(a, 0, 1)))
    np.fill_diagonal(d, np.nan)
    return d

dist      = haversine_matrix(lat, lon)
dist_safe = np.where(np.isnan(dist), 1.0, np.maximum(dist, 50.0))  # 50nm floor

# Pre-compute size vectors (these don't change with k)
so = size ** alpha_o   # origin mass
sd = size ** alpha_d   # destination mass

# ── 4. Steady-state via power iteration ──────────────────────────────────────
def steady_state(P, tol=1e-12, max_iter=200000):
    """Compute steady-state distribution π satisfying πP = π."""
    pi = np.full(P.shape[0], 1.0 / P.shape[0])
    for i in range(max_iter):
        pi_next = pi @ P
        if np.abs(pi_next - pi).sum() < tol:
            return pi_next, i + 1
        pi = pi_next
    return pi_next, max_iter

def build_transition_matrix(k):
    """Build row-stochastic transition matrix for a given distance decay k."""
    dist_term = np.ones((N, N)) if k == 0 else dist_safe ** k
    raw = C * np.outer(so, sd) / dist_term
    np.fill_diagonal(raw, 0)
    return raw / raw.sum(axis=1, keepdims=True)

# ── 5. Sensitivity analysis across k values ──────────────────────────────────
k_values    = [0.0, 0.5, 0.75, 1.0, 1.5]
all_results = {}
ktm_idx     = df[df["iata"] == "KTM"].index[0]

print(f"\n{'k':>6} | {'iters':>6} | {'Spearman ρ':>11} | {'p-value':>10} | "
      f"{'KTM rank':>9} | {'Spread':>7} | Top 5")
print("-" * 95)

for k in k_values:
    P          = build_transition_matrix(k)
    pi, iters  = steady_state(P)
    rho, pval  = spearmanr(pi, size)
    ktm_rank   = int((-pi).argsort().argsort()[ktm_idx] + 1)

    res = df.copy()
    res["centrality"]      = pi
    res["centrality_pct"]  = pi * 100
    res["centrality_rank"] = res["centrality"].rank(ascending=False).astype(int)
    res["h3_util"]         = np.where(res["total"] > 0,
                                       res["h3pax"] / res["total"], 0.0)
    res = res.sort_values("centrality", ascending=False).reset_index(drop=True)
    all_results[k] = res

    spread = res["centrality_pct"].iloc[0] / res["centrality_pct"].iloc[-1]
    top5   = ", ".join(res["iata"].head(5).tolist())
    print(f"k={k:>4} | {iters:>6} | rho={rho:>7.4f}   | p={pval:>8.2e} | "
          f"{ktm_rank:>9} | {spread:>6.1f}x | {top5}")

# ── 6. Rank comparison table (top 20 by k=1.0) ───────────────────────────────
print("\n\nRANK COMPARISON — top 20 airports at k=1.0 across all k values:")
base       = all_results[1.0][["iata", "name", "centrality_rank"]].head(20).copy()
base       = base.rename(columns={"centrality_rank": "rank_k1.0"})
for k in [0.0, 0.5, 0.75, 1.5]:
    lookup = all_results[k].set_index("iata")["centrality_rank"]
    base[f"rank_k{k}"] = base["iata"].map(lookup)
print(base.to_string(index=False))

# ── 7. Illustrative 13×13 submatrix (k=1.0, for appendix) ───────────────────
SUBSET = ["KTM","PBH","GAU","LXA","DEL","SIN","BKK","LHR","ATL","IST","NRT","DOH","JFK"]
idx    = [df[df["iata"] == s].index[0] for s in SUBSET]
P_k1   = build_transition_matrix(1.0)
P_sub  = P_k1[np.ix_(idx, idx)]
df_sub = pd.DataFrame(P_sub * 1000, index=SUBSET, columns=SUBSET).round(3)

print("\n\n13×13 TRANSITION SUBMATRIX at k=1.0 (values ×10⁻³)")
print("Note: rows do not sum to 1 — remaining probability mass flows to")
print("the other 116 airports not shown in this excerpt.\n")
print(df_sub.to_string())

# ── 8. Save all results to Excel ─────────────────────────────────────────────
with pd.ExcelWriter("centrality_sensitivity.xlsx") as writer:
    for k in k_values:
        sheet = all_results[k][["iata", "name", "centrality_pct", "centrality_rank",
                                  "total", "h3pax", "h3_util", "lat", "lon"]]
        sheet.to_excel(writer, sheet_name=f"k={k}", index=False)
    base.to_excel(writer, sheet_name="rank_comparison", index=False)
    df_sub.to_excel(writer, sheet_name="submatrix_k1.0")

print("\nSaved centrality_sensitivity.xlsx")